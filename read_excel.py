import openpyxl
import sys

wb = openpyxl.load_workbook(r'E:\Desktop\Mess management\Alif-Deep Axis(March)-2026.xlsx', data_only=False)

with open(r'E:\Desktop\Mess management\excel_dump.txt', 'w', encoding='utf-8') as f:
    f.write(f'Sheets: {wb.sheetnames}\n')
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f'\n\n{"="*60}\n')
        f.write(f'Sheet: {sheet_name} | rows={ws.max_row} cols={ws.max_column}\n')
        f.write(f'{"="*60}\n')
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    f.write(f'  {cell.coordinate}: {repr(cell.value)}\n')

print('Done. Written to excel_dump.txt')
